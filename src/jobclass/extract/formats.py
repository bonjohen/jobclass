"""Format conversion: ZIP extraction, XLSX-to-CSV/TSV conversion.

Bridges the gap between raw downloaded artifacts (ZIP, XLSX) and parsers
that expect CSV or TSV text.
"""

from __future__ import annotations

import csv
import io
import zipfile

import openpyxl


def extract_xlsx_from_zip(data: bytes) -> bytes:
    """Extract the first .xlsx file from a ZIP archive.

    BLS distributes OEWS data as ZIP files containing a single XLSX workbook.
    Returns the raw XLSX bytes.
    """
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        xlsx_names = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
        if not xlsx_names:
            raise ValueError(f"No .xlsx file found in ZIP. Contents: {zf.namelist()}")
        return zf.read(xlsx_names[0])


def _find_header_row(ws, max_scan: int = 20) -> int:
    """Find the first row that looks like a data header (multiple non-None cells).

    BLS XLSX files often have multi-row preambles (agency name, date, notes)
    before the actual column headers. Returns 1-based row index.
    """
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        non_none = [c for c in row if c is not None]
        # A header row typically has multiple columns populated
        if len(non_none) >= 3:
            return i
    return 1


def xlsx_to_csv(data: bytes, sheet_name: str | None = None, skip_preamble: bool = True) -> str:
    """Convert an XLSX workbook to CSV text.

    If skip_preamble is True, scans for the first row with multiple populated
    cells and treats it as the header. Used for OEWS and SOC files.
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    start_row = _find_header_row(ws) if skip_preamble else 1

    output = io.StringIO()
    writer = csv.writer(output)
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        writer.writerow(row)
    wb.close()

    return output.getvalue()


def xlsx_to_tsv(data: bytes, sheet_name: str | None = None, skip_preamble: bool = True) -> str:
    """Convert an XLSX workbook to TSV text.

    If skip_preamble is True, scans for the first row with multiple populated
    cells and treats it as the header.
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    start_row = _find_header_row(ws) if skip_preamble else 1

    output = io.StringIO()
    writer = csv.writer(output, delimiter="\t")
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        writer.writerow(row)
    wb.close()

    return output.getvalue()


def convert_to_text(data: bytes, expected_format: str, sheet_name: str | None = None) -> str:
    """Dispatch format conversion based on manifest expected_format field.

    Supported formats:
        csv         — decode UTF-8 (no conversion needed)
        tsv         — decode UTF-8 (no conversion needed)
        xlsx_in_zip — extract XLSX from ZIP, then convert to CSV
        xlsx        — convert XLSX to TSV (or CSV depending on parser needs)
    """
    if expected_format == "csv" or expected_format == "tsv":
        return data.decode("utf-8-sig")
    elif expected_format == "xlsx_in_zip":
        xlsx_bytes = extract_xlsx_from_zip(data)
        return xlsx_to_csv(xlsx_bytes)
    elif expected_format == "xlsx":
        return xlsx_to_csv(data, sheet_name=sheet_name)
    else:
        raise ValueError(f"Unsupported format: {expected_format}")
