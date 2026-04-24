"""CPI domain expansion parsers.

Parses BLS CPI flat files (tab-separated) for the full CPI analytical domain:
item hierarchy, area definitions, series metadata, observations, relative
importance, and average prices.

The existing cpi.py handles the single-series deflator (CUSR0000SA0 annual
averages). This module handles the full domain.
"""

from __future__ import annotations

from dataclasses import dataclass

PARSER_VERSION = "2.0.0"

# ---------------------------------------------------------------------------
# Hierarchy level mapping: display_level (from cu.item) → human label
# ---------------------------------------------------------------------------
ITEM_HIERARCHY_LEVELS = {
    "0": "All items",
    "1": "Major group",
    "2": "Intermediate aggregate",
    "3": "Expenditure class",
    "4": "Item stratum",
    "5": "ELI",
    "6": "ELI sub-item",
    "7": "ELI sub-item",
    "8": "ELI sub-item",
}

# Area type classification by area_code pattern
AREA_TYPE_RULES = {
    "0000": "national",
    # 0100-0400 = regions, 01xx/02xx/03xx/04xx sub = divisions
    # Axxx = size class, Dxxx = size class, Sxxx = metro/cross-class
}

# Cross-cutting item codes: these are special aggregates, not strict tree nodes
_CROSS_CUTTING_PREFIXES = ("SA0L", "SA0E", "SAE", "SAC", "SAS", "SAR")
_CROSS_CUTTING_CODES = {"SA0R"}  # Purchasing power of consumer dollar
_AVERAGE_PRICE_PREFIXES = ("SS",)


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------


@dataclass
class CpiItemRow:
    """Parsed CPI item from cu.item."""

    item_code: str
    item_name: str
    hierarchy_level: str
    display_level: int
    selectable: bool
    sort_sequence: int
    parent_item_code: str | None
    source_release_id: str
    parser_version: str


@dataclass
class CpiAreaRow:
    """Parsed CPI area from cu.area."""

    area_code: str
    area_title: str
    area_type: str
    display_level: int
    selectable: bool
    sort_sequence: int
    source_release_id: str
    parser_version: str


@dataclass
class CpiSeriesRow:
    """Parsed CPI series metadata from cu.series."""

    series_id: str
    area_code: str
    item_code: str
    seasonal_adjustment: str  # S or U
    periodicity: str  # R (monthly) or S (semi-annual)
    base_code: str
    base_period: str
    series_title: str
    begin_year: int | None
    end_year: int | None
    source_release_id: str
    parser_version: str


@dataclass
class CpiObservationRow:
    """Parsed CPI index observation from cu.data files."""

    series_id: str
    year: int
    period: str
    value: float
    footnote_codes: str
    source_release_id: str
    parser_version: str


@dataclass
class CpiRelativeImportanceRow:
    """Parsed CPI relative importance entry."""

    item_code: str
    area_code: str
    reference_period: str
    relative_importance: float
    source_release_id: str
    parser_version: str


@dataclass
class CpiAveragePriceRow:
    """Parsed CPI average price observation from ap.data files."""

    series_id: str
    item_code: str
    area_code: str
    year: int
    period: str
    average_price: float
    footnote_codes: str
    source_release_id: str
    parser_version: str


@dataclass
class CpiOverlaySeriesRow:
    """Parsed community overlay series entry (Cleveland Fed, FRED).

    Overlay members are loaded into dim_cpi_member with semantic_role = 'external_overlay'
    and linked to SA0 via bridge_cpi_member_relation.
    """

    member_code: str  # e.g. "MEDCPI", "TRMCPI"
    title: str
    overlay_source: str  # e.g. "cleveland_fed", "fred"
    year: int
    period: str
    value: float
    source_release_id: str
    parser_version: str


@dataclass
class CpiRevisionVintageRow:
    """Parsed C-CPI-U revision vintage entry.

    Tracks preliminary vs. revised index values for the same member/area/period.
    C-CPI-U is initially published as a preliminary estimate, then revised over
    roughly 10-12 months as final expenditure data become available.
    """

    item_code: str
    area_code: str
    year: int
    period: str
    vintage_label: str  # e.g. "2024-preliminary", "2024-final"
    index_value: float
    is_preliminary: bool
    source_release_id: str
    parser_version: str


# ---------------------------------------------------------------------------
# Area type classifier
# ---------------------------------------------------------------------------


def _classify_area_type(area_code: str) -> str:
    """Classify a BLS CPI area code into a type category.

    BLS area code patterns:
    - 0000 = national (U.S. city average)
    - 0100-0400 = regions (Northeast, Midwest, South, West)
    - 01xx-04xx with non-zero suffix = divisions
    - Axxx, Dxxx = size classes
    - Sxxx where all digits (S100-S400) = cross-classification (region × size)
    - Sxxx with letter suffix (S12A, S35A, S49E) = metro areas
    """
    if area_code == "0000":
        return "national"
    c = area_code[0]
    if c == "0":
        # 0100-0400 are regions; 01xx with xx>0 are divisions
        sub = int(area_code[2:])
        if sub == 0:
            return "region"
        return "division"
    if c in ("A", "D"):
        return "size_class"
    if c == "S":
        # Pure numeric Sxxx (e.g., S100, S300) = cross-classification
        # Codes with any letter after S (e.g., S12A, S35A, S49E) = metro
        if area_code[1:].isdigit():
            return "cross_classification"
        return "metro"
    return "other"


def _classify_publication_frequency(area_code: str) -> str:
    """Determine publication frequency for a CPI area."""
    # National and region-level areas are monthly
    if area_code == "0000":
        return "monthly"
    c = area_code[0]
    if c == "0":
        return "monthly"
    if c in ("A", "D"):
        return "monthly"
    # Metro areas: some monthly, some bimonthly
    # Size Class A metros are monthly, smaller metros bimonthly
    if c == "S":
        second = area_code[1]
        if second.isdigit():
            cls = int(second)
            if cls <= 2:
                return "monthly"
            return "bimonthly"
        # Named metros with letter codes — check if Size A
        if second in ("1", "2", "3", "4"):
            return "monthly"
        return "bimonthly"
    return "monthly"


# ---------------------------------------------------------------------------
# Item hierarchy: parent inference
# ---------------------------------------------------------------------------


def _infer_parents(items: list[tuple[str, int]]) -> dict[str, str | None]:
    """Infer parent item codes from display_level ordering.

    BLS cu.item is sorted by sort_sequence with display_level indicating depth.
    The parent of an item at level N is the most recent item at level N-1.
    """
    parent_map: dict[str, str | None] = {}
    stack: list[tuple[str, int]] = []  # (item_code, level)

    for code, level in items:
        # Pop stack until we find the parent level
        while stack and stack[-1][1] >= level:
            stack.pop()
        parent_map[code] = stack[-1][0] if stack else None
        stack.append((code, level))

    return parent_map


# ---------------------------------------------------------------------------
# Semantic role classifier
# ---------------------------------------------------------------------------


def _classify_semantic_role(item_code: str) -> str:
    """Classify a CPI item's semantic role."""
    if item_code in _CROSS_CUTTING_CODES:
        return "purchasing_power"
    for prefix in _CROSS_CUTTING_PREFIXES:
        if item_code.startswith(prefix):
            return "special_aggregate"
    for prefix in _AVERAGE_PRICE_PREFIXES:
        if item_code.startswith(prefix):
            return "average_price_item"
    return "hierarchy_node"


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------


def parse_cpi_item_hierarchy(
    content: str,
    source_release_id: str,
) -> list[CpiItemRow]:
    """Parse BLS cu.item file into item rows with inferred parent-child edges.

    File format: tab-separated with columns
    item_code, item_name, display_level, selectable, sort_sequence
    """
    rows: list[CpiItemRow] = []
    items_for_parents: list[tuple[str, int]] = []

    for line in content.splitlines():
        line = line.rstrip()
        if not line or line.startswith("item_code"):
            continue
        parts = line.split("\t")
        if len(parts) < 5:
            continue
        code = parts[0].strip()
        name = parts[1].strip()
        try:
            display_level = int(parts[2].strip())
        except ValueError:
            continue
        selectable = parts[3].strip().upper() == "T"
        try:
            sort_seq = int(parts[4].strip())
        except ValueError:
            sort_seq = 0

        level_label = ITEM_HIERARCHY_LEVELS.get(str(display_level), f"Level {display_level}")
        items_for_parents.append((code, display_level))
        rows.append(
            CpiItemRow(
                item_code=code,
                item_name=name,
                hierarchy_level=level_label,
                display_level=display_level,
                selectable=selectable,
                sort_sequence=sort_seq,
                parent_item_code=None,  # filled below
                source_release_id=source_release_id,
                parser_version=PARSER_VERSION,
            )
        )

    # Infer parent-child relationships from display_level ordering
    parent_map = _infer_parents(items_for_parents)
    for row in rows:
        row.parent_item_code = parent_map.get(row.item_code)

    return rows


def parse_cpi_area(
    content: str,
    source_release_id: str,
) -> list[CpiAreaRow]:
    """Parse BLS cu.area file into area rows.

    File format: tab-separated with columns
    area_code, area_name, display_level, selectable, sort_sequence
    """
    rows: list[CpiAreaRow] = []
    for line in content.splitlines():
        line = line.rstrip()
        if not line or line.startswith("area_code"):
            continue
        parts = line.split("\t")
        if len(parts) < 5:
            continue
        code = parts[0].strip()
        name = parts[1].strip()
        try:
            display_level = int(parts[2].strip())
        except ValueError:
            continue
        selectable = parts[3].strip().upper() == "T"
        try:
            sort_seq = int(parts[4].strip())
        except ValueError:
            sort_seq = 0

        rows.append(
            CpiAreaRow(
                area_code=code,
                area_title=name,
                area_type=_classify_area_type(code),
                display_level=display_level,
                selectable=selectable,
                sort_sequence=sort_seq,
                source_release_id=source_release_id,
                parser_version=PARSER_VERSION,
            )
        )
    return rows


def parse_cpi_series(
    content: str,
    source_release_id: str,
) -> list[CpiSeriesRow]:
    """Parse BLS cu.series file into series metadata rows.

    File format: tab-separated with columns
    series_id, area_code, item_code, seasonal, periodicity_code, base_code,
    base_period, series_title, footnote_codes, begin_year, begin_period,
    end_year, end_period
    """
    rows: list[CpiSeriesRow] = []
    for line in content.splitlines():
        line = line.rstrip()
        if not line or line.startswith("series_id"):
            continue
        parts = line.split("\t")
        if len(parts) < 8:
            continue
        series_id = parts[0].strip()
        area_code = parts[1].strip()
        item_code = parts[2].strip()
        seasonal = parts[3].strip()
        periodicity = parts[4].strip()
        base_code = parts[5].strip()
        base_period = parts[6].strip()
        series_title = parts[7].strip()

        begin_year = None
        end_year = None
        try:
            if len(parts) > 9:
                begin_year = int(parts[9].strip())
            if len(parts) > 11:
                end_year = int(parts[11].strip())
        except ValueError:
            pass

        rows.append(
            CpiSeriesRow(
                series_id=series_id,
                area_code=area_code,
                item_code=item_code,
                seasonal_adjustment=seasonal,
                periodicity=periodicity,
                base_code=base_code,
                base_period=base_period,
                series_title=series_title,
                begin_year=begin_year,
                end_year=end_year,
                source_release_id=source_release_id,
                parser_version=PARSER_VERSION,
            )
        )
    return rows


def parse_cpi_observations(
    content: str,
    source_release_id: str,
) -> list[CpiObservationRow]:
    """Parse BLS cu.data flat file into observation rows.

    File format: tab-separated with columns
    series_id, year, period, value, footnote_codes

    Unlike the existing cpi.py deflator parser, this parses ALL series and
    ALL periods (not just M13 annual averages).
    """
    rows: list[CpiObservationRow] = []
    for line in content.splitlines():
        line = line.rstrip()
        if not line or line.startswith("series_id"):
            continue
        parts = line.split("\t")
        if len(parts) < 4:
            continue
        series_id = parts[0].strip()
        footnotes = parts[4].strip() if len(parts) > 4 else ""
        try:
            year = int(parts[1].strip())
            value = float(parts[3].strip())
        except (ValueError, IndexError):
            continue
        period = parts[2].strip()

        rows.append(
            CpiObservationRow(
                series_id=series_id,
                year=year,
                period=period,
                value=value,
                footnote_codes=footnotes,
                source_release_id=source_release_id,
                parser_version=PARSER_VERSION,
            )
        )
    return rows


def parse_cpi_relative_importance(
    content: bytes,
    source_release_id: str,
) -> list[CpiRelativeImportanceRow]:
    """Parse BLS relative importance XLSX table.

    BLS publishes relative importance as Excel files with item × area columns.
    The exact layout varies by year, but typically has item codes/names in the
    first columns and area-specific importance values in subsequent columns.
    """
    from io import BytesIO

    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    rows: list[CpiRelativeImportanceRow] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 2:
            continue

        # Find header row — look for "Item" or "item_code" pattern
        header_idx = None
        for i, row in enumerate(all_rows):
            if row and any(
                str(cell).strip().lower() in ("item", "item code", "expenditure category")
                for cell in row
                if cell is not None
            ):
                header_idx = i
                break

        if header_idx is None:
            continue

        header = [str(c).strip() if c else "" for c in all_rows[header_idx]]

        # Find item code column and area columns
        item_col = None
        area_cols: list[tuple[int, str]] = []
        for j, h in enumerate(header):
            hl = h.lower()
            if hl in ("item", "item code", "item_code"):
                item_col = j
            elif hl in ("expenditure category",):
                # Name column, skip — item code should be elsewhere
                continue
            elif h and j > 0:
                # Remaining columns are area-specific importance values
                area_cols.append((j, h))

        if item_col is None:
            # Try first column as item code
            item_col = 0
            area_cols = [(j, h) for j, h in enumerate(header) if j > 0 and h]

        reference_period = sheet_name.strip()

        for data_row in all_rows[header_idx + 1 :]:
            if not data_row or data_row[item_col] is None:
                continue
            item_code = str(data_row[item_col]).strip()
            if not item_code:
                continue

            for col_idx, _area_label in area_cols:
                if col_idx >= len(data_row) or data_row[col_idx] is None:
                    continue
                try:
                    importance = float(data_row[col_idx])
                except (ValueError, TypeError):
                    continue

                # Derive area_code from area_label — default to national
                area_code = "0000"
                rows.append(
                    CpiRelativeImportanceRow(
                        item_code=item_code,
                        area_code=area_code,
                        reference_period=reference_period,
                        relative_importance=importance,
                        source_release_id=source_release_id,
                        parser_version=PARSER_VERSION,
                    )
                )

    wb.close()
    return rows


def parse_cpi_average_prices(
    content: str,
    source_release_id: str,
) -> list[CpiAveragePriceRow]:
    """Parse BLS average price flat file (ap.data format).

    File format: tab-separated with columns
    series_id, year, period, value, footnote_codes

    Average price series IDs follow format: APU{area}{item}
    """
    rows: list[CpiAveragePriceRow] = []
    for line in content.splitlines():
        line = line.rstrip()
        if not line or line.startswith("series_id"):
            continue
        parts = line.split("\t")
        if len(parts) < 4:
            continue
        series_id = parts[0].strip()
        footnotes = parts[4].strip() if len(parts) > 4 else ""

        # Decompose AP series ID: APU + area_code(4) + item_code
        if not series_id.startswith("AP"):
            continue
        # Format: AP + seasonal(1) + area(4) + item(variable)
        if len(series_id) < 8:
            continue
        area_code = series_id[3:7]
        item_code = series_id[7:]

        try:
            year = int(parts[1].strip())
            price = float(parts[3].strip())
        except (ValueError, IndexError):
            continue
        period = parts[2].strip()

        rows.append(
            CpiAveragePriceRow(
                series_id=series_id,
                item_code=item_code,
                area_code=area_code,
                year=year,
                period=period,
                average_price=price,
                footnote_codes=footnotes,
                source_release_id=source_release_id,
                parser_version=PARSER_VERSION,
            )
        )
    return rows


# ---------------------------------------------------------------------------
# Series ID validation
# ---------------------------------------------------------------------------


def validate_series_decomposition(
    series_rows: list[CpiSeriesRow],
    known_item_codes: set[str],
    known_area_codes: set[str],
) -> list[str]:
    """Validate that series ID components match known members and areas.

    Returns list of warning messages for unresolved codes.
    """
    warnings: list[str] = []
    for s in series_rows:
        if s.item_code not in known_item_codes:
            warnings.append(f"Series {s.series_id}: item_code {s.item_code!r} not in known items")
        if s.area_code not in known_area_codes:
            warnings.append(f"Series {s.series_id}: area_code {s.area_code!r} not in known areas")
    return warnings
